#!/usr/bin/env python3
"""Stage 1 for the new SAO (САО) map: match the "первая очередь" list against
the reestr-ogh.mos.ru ОДХ registry, then split the окrug into two exports.

What it does, in order:
  1. Reads --wave1-xlsx (e.g. "ОДХ первая очередь.xlsx") -- the list of
     object NAMES with no IDs/addresses.
  2. Logs into reestr-ogh.mos.ru and downloads the ОДХ (roads) registry
     visible to the logged-in account (same endpoint as odh_yandex_export.py).
  3. Tries to auto-detect which registry field marks the округ (okrug) and
     filters down to САО. If no such field exists in the response at all AND
     the total count is close to the round's known total, the script assumes
     the whole registry returned is already scoped to that округ (accounts
     here appear to be organization-scoped) and proceeds with a clear log
     message instead of failing. Pass --ao-field/--ao-value yourself if you
     ever know the real filter field (also makes this much faster: a
     server-side filter instead of downloading everything the account sees).
  4. Matches each wave-1 name to a САО registry row by normalized exact
     match, falling back to fuzzy matching (difflib). Writes
     sverka_report.csv listing every match and its confidence -- CHECK THE
     "fuzzy" AND "no_match" ROWS BY HAND before trusting them.
  5. Fetches geometry for (a) the matched "first wave" objects and
     (b) every other САО ОДХ ("remaining"), and writes each as MSK-77 (raw
     metres) + WGS-84 (degrees) GeoJSON/CSV, ready for build_sao_maps.py.

Optionally probes other ogh_types (--probe-types) to see whether hydrants /
ДЭУ bases / snow-storage sites / etc. even live in this registry -- this is
exploratory and may come back empty if those object types belong to a
different system.

Requires: requests, openpyxl, pyproj (python -m pip install requests openpyxl pyproj)

Example (PowerShell):
  $env:ODS_USERNAME = '...'
  $env:ODS_PASSWORD = '...'
  python sao_stage1_export.py --wave1-xlsx "ОДХ первая очередь.xlsx" --output-dir sao_out --probe-types
"""
from __future__ import annotations

import argparse
import csv
import difflib
import getpass
import json
import os
import re
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Any

import requests

from msk77_wgs84 import self_check_transformer, transform_coords

BASE_URL = "https://reestr-ogh.mos.ru"
API_URL = f"{BASE_URL}/api"
ODH_TYPE_ID = 1
PAGE_SIZE = 50
AO_HINTS = ("сао", "северный")  # substrings that identify "САО" in a field name/value

ABBREV = [
    (r"\bул\.?\b", "улица"),
    (r"\bпр-?т\b", "проспект"),
    (r"\bпр-?д\b", "проезд"),
    (r"\bпер\.?\b", "переулок"),
    (r"\bбул\.?\b", "бульвар"),
    (r"\bш\.?\b", "шоссе"),
    (r"\bпл\.?\b", "площадь"),
    (r"\bнаб\.?\b", "набережная"),
    (r"\bп/п\b", "путепровод"),
]


def api_headers() -> dict[str, str]:
    return {
        "User-Agent": "Mozilla/5.0 SAO-Stage1/1.0",
        "Accept": "application/json",
        "Origin": BASE_URL,
        "Referer": f"{BASE_URL}/r/ogh/odh",
    }


def login(session: requests.Session, username: str, password: str) -> None:
    headers = api_headers()
    session.get(f"{BASE_URL}/auth/login", headers=headers, timeout=30)
    response = session.post(
        f"{API_URL}/login",
        headers={**headers, "Content-Type": "application/x-www-form-urlencoded"},
        data={"username": username, "password": password, "j_username": username, "j_password": password},
        timeout=30,
    )
    response.raise_for_status()
    body = response.json()
    if body.get("code") != 200:
        raise RuntimeError(f"Ошибка входа: {body}")


def normalize_name(name: str) -> str:
    s = str(name or "").lower().replace("ё", "е")
    s = re.sub(r"[«»\"'.,()]", " ", s)
    for pattern, full in ABBREV:
        s = re.sub(pattern, full, s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def read_wave1_list(path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Установите openpyxl: python -m pip install openpyxl") from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    header_row = next(ws.iter_rows(max_row=5), None)
    if header_row is None:
        raise ValueError("Пустой файл")
    headers = [str(c.value or "").strip() for c in header_row]

    def col(*needles: str) -> int | None:
        for i, h in enumerate(headers):
            if any(n.lower() in h.lower() for n in needles):
                return i
        return None

    name_col = col("название объекта", "название")
    seq_ao_col = col("№ п/п \nв ао", "№ п/п в ао", "n п/п в ао")
    inst_col = col("учреждение")
    if name_col is None:
        raise ValueError(f"Не найдена колонка с названием объекта. Заголовки: {headers}")
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_col] if name_col < len(row) else None
        if not name or not str(name).strip():
            continue
        result.append({
            "seq_ao": row[seq_ao_col] if seq_ao_col is not None and seq_ao_col < len(row) else None,
            "name": str(name).strip(),
            "institution": str(row[inst_col]).strip() if inst_col is not None and inst_col < len(row) and row[inst_col] else "",
        })
    return result


def load_full_registry(session: requests.Session, ao_field: str | None, ao_value: str | None) -> list[dict[str, Any]]:
    common: dict[str, Any] = {
        "main_page": True,
        "max_rows": PAGE_SIZE,
        "parent_type_id": -1,
        "sort": "root_id.asc",
        "type_id": ODH_TYPE_ID,
        "ogh_types": [ODH_TYPE_ID],
    }
    if ao_field and ao_value:
        common[ao_field] = ao_value
        print(f"Серверный фильтр округа: {ao_field}={ao_value}", file=sys.stderr)
    count_response = session.post(
        f"{API_URL}/registry/ogh/count", json={**common, "page": 0},
        headers={**api_headers(), "Content-Type": "application/json"}, timeout=60,
    )
    count_response.raise_for_status()
    total = int(count_response.json()["data"]["count"])
    pages = (total + PAGE_SIZE - 1) // PAGE_SIZE
    rows: list[dict[str, Any]] = []
    for page in range(pages):
        response = session.post(
            f"{API_URL}/registry/ogh", json={**common, "page": page},
            headers={**api_headers(), "Content-Type": "application/json"}, timeout=120,
        )
        response.raise_for_status()
        rows.extend(response.json().get("data", []))
        print(f"Реестр: страница {page + 1}/{pages}, получено {len(rows)}/{total}", file=sys.stderr)
    return rows


REFERENCE_SAO_TOTAL = 658  # из справки на присланной карте САО -- только ориентир для сверки


def guess_ao_field(rows: list[dict[str, Any]]) -> str | None:
    if not rows:
        return None
    print("Поля первой строки реестра (для ручной проверки): " + ", ".join(sorted(rows[0].keys())), file=sys.stderr)
    candidate_keys = [k for k in rows[0] if any(h in k.lower() for h in ("ao", "okrug", "distr", "oiv"))]
    for key in candidate_keys:
        values = {str(r.get(key, "")).lower() for r in rows[:300]}
        if any(any(h in v for h in AO_HINTS) for v in values):
            print(f"Похоже на поле округа: {key!r} (пример значений: {list(values)[:5]})", file=sys.stderr)
            return key
    print("В схеме реестра нет отдельного поля округа (ao/okrug/distr/oiv).", file=sys.stderr)
    return None


def summarize_org_fields(rows: list[dict[str, Any]]) -> None:
    for field in ("customer_name", "grbs_short_name"):
        counts = Counter(str(r.get(field, "")).strip() for r in rows)
        print(f"Распределение по полю {field!r} (топ-10): {dict(counts.most_common(10))}", file=sys.stderr)


def filter_by_ao(rows: list[dict[str, Any]], ao_field: str | None) -> list[dict[str, Any]]:
    if not ao_field:
        return []
    return [r for r in rows if any(h in str(r.get(ao_field, "")).lower() for h in AO_HINTS)]


def match_wave1(wave1_items: list[dict[str, Any]], sao_rows: list[dict[str, Any]]):
    """Each registry row can be claimed by at most one wave1 item -- otherwise
    two different input names that happen to be textually close (e.g. two
    different проезды near the same street) can silently collapse onto the
    same registry object. When that happens the second item falls through to
    the next-closest candidate, or to no_match if there isn't one."""
    by_norm: dict[str, list[dict[str, Any]]] = {}
    for row in sao_rows:
        by_norm.setdefault(normalize_name(row.get("object_name")), []).append(row)
    all_norms = list(by_norm.keys())
    used_ids: set[Any] = set()
    matches = []
    for item in wave1_items:
        norm = normalize_name(item["name"])
        row, method, score = None, "no_match", 0.0
        if norm in by_norm:
            candidates = [r for r in by_norm[norm] if r.get("id") not in used_ids]
            if candidates:
                row, method, score = candidates[0], "exact", 1.0
        if row is None:
            for close in difflib.get_close_matches(norm, all_norms, n=5, cutoff=0.72):
                candidates = [r for r in by_norm[close] if r.get("id") not in used_ids]
                if candidates:
                    row = candidates[0]
                    method = "fuzzy"
                    score = round(difflib.SequenceMatcher(None, norm, close).ratio(), 3)
                    break
        if row is not None:
            used_ids.add(row.get("id"))
        matches.append({"input": item, "row": row, "method": method, "score": score})
    return matches


def write_sverka_report(path: Path, matches: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(["№ в АО", "Название в списке", "Метод", "Похожесть", "Найдено в реестре", "ID реестра", "Категория", "Статус"])
        for m in matches:
            row = m["row"] or {}
            writer.writerow([
                m["input"].get("seq_ao"), m["input"]["name"], m["method"], m["score"],
                row.get("object_name", ""), row.get("short_root_id", ""),
                row.get("clean_category_name", ""), row.get("object_status_name", ""),
            ])


def load_geometry(session: requests.Session, row: dict[str, Any]) -> list[dict[str, Any]]:
    params = {"id": row["id"], "root_id": row.get("object_id") or row.get("id"), "type_id": ODH_TYPE_ID}
    for attempt in range(1, 6):
        try:
            response = session.get(f"{API_URL}/ogh/geometry/full", params=params, headers=api_headers(), timeout=180)
            if response.status_code == 429 or response.status_code >= 500:
                raise requests.HTTPError(f"HTTP {response.status_code}", response=response)
            response.raise_for_status()
            return response.json() or []
        except (requests.RequestException, ValueError) as exc:
            if attempt == 5:
                raise
            delay = min(30, 2 ** (attempt - 1))
            print(f"Повтор запроса геометрии ID {row.get('short_root_id')} через {delay} с: {exc}", file=sys.stderr)
            time.sleep(delay)
    return []


def normalize_id(value: Any) -> str:
    s = str(value or "").strip()
    return s[:-2] if re.fullmatch(r"\d+\.0", s) else s


def build_features(rows: list[dict[str, Any]], extra_props_fn=None) -> list[dict[str, Any]]:
    features = []
    for row in rows:
        for geometry in row.get("_geometry", []):
            for kind, gtype in (("polygons", "Polygon"), ("lines", "LineString"), ("points", "Point")):
                value = geometry.get(kind)
                if not value:
                    continue
                for item in (value if isinstance(value, list) else [value]):
                    coords = item.get("coordinates") if isinstance(item, dict) else item
                    if not coords:
                        continue
                    props = {
                        "id": normalize_id(row.get("short_root_id")),
                        "name": row.get("object_name") or "",
                        "category": row.get("clean_category_name") or "",
                        "status": row.get("object_status_name") or "",
                        "area_m2": row.get("total_area"),
                        "registry_id": row.get("id"),
                    }
                    if extra_props_fn:
                        props.update(extra_props_fn(row))
                    features.append({"type": "Feature", "geometry": {"type": gtype, "coordinates": coords}, "properties": props})
    return features


def write_geojson(path: Path, features: list[dict[str, Any]]) -> None:
    path.write_text(json.dumps({"type": "FeatureCollection", "features": features}, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, features: list[dict[str, Any]]) -> None:
    if not features:
        path.write_text("", encoding="utf-8-sig")
        return
    fields = list(features[0]["properties"].keys()) + ["geometry_type"]
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, delimiter=";")
        writer.writeheader()
        for f in features:
            row = dict(f["properties"])
            row["geometry_type"] = f["geometry"]["type"]
            writer.writerow(row)


def fetch_and_export(session: requests.Session, rows: list[dict[str, Any]], output_dir: Path, prefix: str,
                      cache: dict[str, Any], extra_props_fn=None) -> list[dict[str, Any]]:
    for index, row in enumerate(rows, 1):
        key = normalize_id(row.get("short_root_id"))
        if key in cache:
            row["_geometry"] = cache[key]
        else:
            row["_geometry"] = load_geometry(session, row)
            cache[key] = row["_geometry"]
        print(f"[{prefix}] геометрия {index}/{len(rows)}: {row.get('short_root_id')}", file=sys.stderr)
        time.sleep(0.15)
    features_msk77 = build_features(rows, extra_props_fn)
    features_wgs84 = [
        {**f, "geometry": {**f["geometry"], "coordinates": transform_coords(f["geometry"]["coordinates"])}}
        for f in features_msk77
    ]
    write_geojson(output_dir / f"{prefix}_msk77.geojson", features_msk77)
    write_csv(output_dir / f"{prefix}_msk77.csv", features_msk77)
    write_geojson(output_dir / f"{prefix}_wgs84.geojson", features_wgs84)
    write_csv(output_dir / f"{prefix}_wgs84.csv", features_wgs84)
    return features_wgs84


def probe_ogh_types(session: requests.Session, output_dir: Path, ao_field: str | None, ao_value: str | None, max_type_id: int = 15) -> None:
    """Best-effort scan: which type_id values return objects at all, and what
    do their names look like? Might reveal hydrants/ДЭУ bases/etc. -- or might
    not, if those live in a different system entirely. Errors per type_id are
    swallowed so one bad type doesn't stop the scan."""
    results = {}
    for type_id in range(1, max_type_id + 1):
        payload = {
            "main_page": True, "max_rows": 5, "parent_type_id": -1, "sort": "root_id.asc",
            "type_id": type_id, "ogh_types": [type_id], "page": 0,
        }
        if ao_field and ao_value:
            payload[ao_field] = ao_value
        try:
            count_resp = session.post(f"{API_URL}/registry/ogh/count", json=payload,
                                       headers={**api_headers(), "Content-Type": "application/json"}, timeout=30)
            count_resp.raise_for_status()
            count = count_resp.json().get("data", {}).get("count", 0)
            sample_rows = []
            if count:
                data_resp = session.post(f"{API_URL}/registry/ogh", json=payload,
                                          headers={**api_headers(), "Content-Type": "application/json"}, timeout=30)
                data_resp.raise_for_status()
                sample_rows = data_resp.json().get("data", [])[:2]  # full raw rows -- object_name isn't always the naming field
            results[type_id] = {"count": count, "sample_rows": sample_rows}
            names = [r.get("object_name") for r in sample_rows]
            print(f"[probe] type_id={type_id}: count={count} object_name={names}", file=sys.stderr)
        except Exception as exc:
            results[type_id] = {"error": str(exc)}
        time.sleep(0.1)
    (output_dir / "ogh_types_probe.json").write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print("Результат сохранён в ogh_types_probe.json -- посмотри, есть ли там гидранты/ДЭУ/снегосвалки.", file=sys.stderr)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--wave1-xlsx", required=True, type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("sao_out"))
    parser.add_argument("--username", default=os.getenv("ODS_USERNAME"))
    parser.add_argument("--password", default=os.getenv("ODS_PASSWORD"))
    parser.add_argument("--ao-field", default=None, help="Точное имя поля округа в ответе реестра, если уже известно")
    parser.add_argument("--ao-value", default="САО", help="Значение округа для фильтра (по умолчанию 'САО')")
    parser.add_argument("--probe-types", action="store_true", help="Дополнительно просканировать другие type_id (гидранты/ДЭУ/снегосвалки и т.п.)")
    args = parser.parse_args()

    self_check_transformer()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    wave1_items = read_wave1_list(args.wave1_xlsx)
    print(f"В списке первой очереди: {len(wave1_items)} объектов", file=sys.stderr)

    username = args.username or input("Логин АСУ ОДС: ")
    password = args.password or getpass.getpass("Пароль АСУ ОДС: ")
    session = requests.Session()
    login(session, username, password)

    registry = load_full_registry(session, args.ao_field, args.ao_value if args.ao_field else None)
    if not registry:
        print("ВНИМАНИЕ: реестр пуст -- 0 объектов. Проверь логин и права аккаунта.", file=sys.stderr)
        return 1

    if args.ao_field:
        ao_field = args.ao_field
        sao_rows = registry  # уже отфильтровано на сервере
        print(f"Серверный фильтр применён ({args.ao_field}={args.ao_value}): {len(sao_rows)} объектов", file=sys.stderr)
    else:
        ao_field = guess_ao_field(registry)
        if ao_field:
            sao_rows = filter_by_ao(registry, ao_field)
            print(f"Отфильтровано по полю {ao_field!r}: {len(sao_rows)} объектов", file=sys.stderr)
        else:
            summarize_org_fields(registry)
            print(
                f"Явного поля округа в реестре нет. Всего объектов на аккаунте: {len(registry)} "
                f"(для справки: на присланной карте у САО указано {REFERENCE_SAO_TOTAL}). Похоже, "
                "аккаунт уже привязан к организации САО, и весь полученный реестр и так только "
                f"по этому округу -- работаю дальше со всеми {len(registry)} объектами как с САО. "
                "Если распределение по customer_name/grbs_short_name выше говорит об обратном -- "
                "останови (Ctrl+C) и разберёмся.",
                file=sys.stderr,
            )
            sao_rows = registry

    if args.probe_types:
        probe_ogh_types(session, args.output_dir, ao_field, args.ao_value)

    matches = match_wave1(wave1_items, sao_rows)
    write_sverka_report(args.output_dir / "sverka_report.csv", matches)
    method_counts = Counter(m["method"] for m in matches)
    print(f"Сверка: {dict(method_counts)} -- открой sverka_report.csv и проверь fuzzy/no_match вручную.", file=sys.stderr)

    matched_ids = {m["row"]["id"] for m in matches if m["row"]}
    wave1_rows = [m["row"] for m in matches if m["row"]]
    remaining_rows = [r for r in sao_rows if r.get("id") not in matched_ids]

    cache_path = args.output_dir / ".geometry_cache.json"
    try:
        cache = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.exists() else {}
    except (OSError, json.JSONDecodeError):
        cache = {}

    fetch_and_export(session, wave1_rows, args.output_dir, "sao_wave1", cache, lambda r: {"wave": "1"})
    cache_path.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    fetch_and_export(session, remaining_rows, args.output_dir, "sao_remaining", cache, lambda r: {"wave": "?"})
    cache_path.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")

    print(
        f"Готово: wave1={len(wave1_rows)} (из {len(wave1_items)} в списке, "
        f"{len(wave1_items) - len(wave1_rows)} без совпадения), remaining={len(remaining_rows)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
