#!/usr/bin/env python3
"""Extract PGM-container data and link it to exported ODH road geometry by ASU ODS ID."""
from __future__ import annotations
import argparse,csv,json
from pathlib import Path
from openpyxl import load_workbook

FIELDS=['number','ao','institution','odh_name','odh_id','address_hint','container_count','placement_type','geometry_status']

def representative_point(geometry:dict):
    c=geometry['coordinates']; typ=geometry['type']
    if typ=='Point': return c
    if typ=='LineString': return c[len(c)//2]
    if typ=='Polygon':
        pts=geometry['coordinates'][0]; return pts[len(pts)//2]
    return None

def road_index(data_dir:Path):
    idx={}
    for name in ('sao_wave1_wgs84.geojson','sao_remaining_wgs84.geojson'):
        for f in json.loads((data_dir/name).read_text(encoding='utf-8'))['features']:
            p=f.get('properties') or {}; oid=str(p.get('id',''))
            if oid and oid not in idx and f.get('geometry'): idx[oid]=f
    return idx

def main()->int:
 p=argparse.ArgumentParser();p.add_argument('--input',type=Path,required=True);p.add_argument('--data-dir',type=Path,required=True);p.add_argument('--output-dir',type=Path,required=True);a=p.parse_args();a.output_dir.mkdir(parents=True,exist_ok=True)
 wb=load_workbook(a.input,read_only=True,data_only=True); ws=wb.worksheets[0]; roads=road_index(a.data_dir); rows=[]; features=[]
 for values in ws.iter_rows(min_row=5,values_only=True):
  if not values or values[0] is None: continue
  try: n=str(int(float(values[0])))
  except (TypeError,ValueError): continue
  oid=str(values[4] or '').strip(); linked=roads.get(oid); status='odh_geometry_proxy' if linked else 'no_odh_geometry_match'
  row={'number':n,'ao':str(values[1] or '').strip(),'institution':str(values[2] or '').strip(),'odh_name':str(values[3] or '').strip(),'odh_id':oid,'address_hint':str(values[5] or '').strip(),'container_count':str(values[6] or '').strip(),'placement_type':str(values[7] or '').strip(),'geometry_status':status}; rows.append(row)
  if linked:
   point=representative_point(linked['geometry'])
   if point: features.append({'type':'Feature','geometry':{'type':'Point','coordinates':point},'properties':{**row,'name':f"ПГМ: {row['odh_name']}",'warning':'Точка — представитель геометрии ОДХ; точное место контейнера уточняется адресным ориентиром.'}})
 with (a.output_dir/'sao_pgm_containers.csv').open('w',encoding='utf-8-sig',newline='') as f: w=csv.DictWriter(f,fieldnames=FIELDS,delimiter=';');w.writeheader();w.writerows(rows)
 (a.output_dir/'sao_pgm_containers_wgs84.geojson').write_text(json.dumps({'type':'FeatureCollection','metadata':{'record_count':len(rows),'geometry_proxy_count':len(features),'note':'Точки размещены как представители геометрии соответствующего ОДХ, а не как координаты контейнера.'},'features':features},ensure_ascii=False,indent=2),encoding='utf-8')
 print(f'Записей ПГМ: {len(rows)}; связаны с ОДХ: {len(features)}; без связи: {len(rows)-len(features)}')
 return 0
if __name__=='__main__':raise SystemExit(main())
