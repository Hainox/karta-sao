#!/usr/bin/env python3
"""Extract historical snow-storage address list from the ЖКХ workplan workbook."""
from __future__ import annotations
import argparse,csv,json,re,time,sys
from pathlib import Path
from openpyxl import load_workbook
import requests

FIELDS=['number','district','address','area_m2','note','source_period','geocode_status','latitude','longitude','geocode_label']

def parse_sites(path:Path)->list[dict[str,str]]:
    wb=load_workbook(path,read_only=True,data_only=True)
    ws=next((s for s in wb.worksheets if 'свалки' in ' '.join(str(c.value or '') for row in s.iter_rows(min_row=1,max_row=4) for c in row).lower()),None)
    if ws is None: raise ValueError('Не найден лист с местами временного складирования снега')
    result=[]
    for row in ws.iter_rows(min_row=5,values_only=True):
        if not row or row[0] is None or row[2] is None: continue
        try: number=str(int(float(row[0])))
        except (TypeError,ValueError): continue
        result.append({'number':number,'district':str(row[1] or '').strip(),'address':str(row[2]).strip(),'area_m2':str(row[3] or '').strip(),'note':str(row[4] or '').strip(),'source_period':'2025-2026','geocode_status':'not_checked','latitude':'','longitude':'','geocode_label':''})
    return result

def geocode(sites:list[dict[str,str]],cache:dict[str,dict],enabled:bool,cache_path:Path,max_new:int)->None:
    new_requests=0
    for site in sites:
        query=f"{site['address']}, Москва, Россия"
        if query in cache: candidates=cache[query]
        elif enabled and new_requests < max_new:
            response=requests.get('https://nominatim.openstreetmap.org/search',params={'q':query,'format':'jsonv2','limit':1},headers={'User-Agent':'SAO-map-preparation/1.0'},timeout=30); response.raise_for_status(); candidates=response.json(); cache[query]=candidates; new_requests += 1; cache_path.write_text(json.dumps(cache,ensure_ascii=False),encoding='utf-8'); print('Geocoder: ' + site['address'], file=sys.stderr, flush=True); time.sleep(1.05)
        else: continue
        if candidates:
            hit=candidates[0]; site.update({'geocode_status':'candidate','latitude':hit['lat'],'longitude':hit['lon'],'geocode_label':hit['display_name']})
        else: site['geocode_status']='not_found'

def main()->int:
    p=argparse.ArgumentParser();p.add_argument('--input',type=Path,required=True);p.add_argument('--output-dir',type=Path,required=True);p.add_argument('--geocode',action='store_true');p.add_argument('--max-new',type=int,default=12);a=p.parse_args();a.output_dir.mkdir(parents=True,exist_ok=True)
    sites=parse_sites(a.input); cache_path=a.output_dir/'snow_sites_geocode_cache.json'; cache=json.loads(cache_path.read_text(encoding='utf-8')) if cache_path.exists() else {}; geocode(sites,cache,a.geocode,cache_path,a.max_new); cache_path.write_text(json.dumps(cache,ensure_ascii=False,indent=2),encoding='utf-8')
    with (a.output_dir/'sao_snow_storage_sites.csv').open('w',encoding='utf-8-sig',newline='') as f: w=csv.DictWriter(f,fieldnames=FIELDS,delimiter=';');w.writeheader();w.writerows(sites)
    features=[]
    for s in sites:
        if s['geocode_status']=='candidate': features.append({'type':'Feature','geometry':{'type':'Point','coordinates':[float(s['longitude']),float(s['latitude'])]},'properties':{**s,'name':f"Место временного складирования снега: {s['address']}",'verification':'Кандидат геокодера; сверить перед служебным использованием'}})
    (a.output_dir/'sao_snow_storage_sites_wgs84.geojson').write_text(json.dumps({'type':'FeatureCollection','features':features},ensure_ascii=False,indent=2),encoding='utf-8')
    print(f'Адресов: {len(sites)}, точек-кандидатов: {len(features)}')
    return 0
if __name__=='__main__': raise SystemExit(main())
