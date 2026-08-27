from openpyxl import load_workbook
from pathlib import Path
import json
p=Path(r'C:\Users\root\Downloads\Управление ЖКХ план график.xlsx')
wb=load_workbook(p, read_only=True, data_only=True)
out=[]
for ws in wb.worksheets:
    rows=[]
    for idx,row in enumerate(ws.iter_rows(values_only=True),1):
        vals=[None if v is None else str(v) for v in row]
        if any(v not in (None,'') for v in vals): rows.append({'row':idx,'values':vals})
        if len(rows)>=25: break
    out.append({'title':ws.title, 'max_row':ws.max_row, 'max_column':ws.max_column, 'nonempty_sample':rows})
Path('work/jkh_workbook_structure.json').write_text(json.dumps(out, ensure_ascii=False,indent=2),encoding='utf-8')
