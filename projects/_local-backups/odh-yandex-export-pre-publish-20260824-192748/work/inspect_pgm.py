from openpyxl import load_workbook
from pathlib import Path
import json
p=Path(r'C:\Users\root\Downloads\Контейнеры для хранения аварийного запаса ПГМ 451 (1).xlsx')
wb=load_workbook(p,read_only=True,data_only=True)
out=[]
for ws in wb.worksheets:
 rows=[]
 for idx,row in enumerate(ws.iter_rows(values_only=True),1):
  vals=[None if v is None else str(v) for v in row]
  if any(v not in (None,'') for v in vals): rows.append({'row':idx,'values':vals})
  if len(rows)>=35: break
 out.append({'title':ws.title,'sample':rows})
Path('work/pgm_containers_structure.json').write_text(json.dumps(out,ensure_ascii=False,indent=2),encoding='utf-8')
