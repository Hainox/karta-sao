from __future__ import annotations

import argparse
import json
import re
from datetime import date
from pathlib import Path

import openpyxl


def clean(value: object) -> str:
    return str(value or "").strip()


def parse_coordinates(value: object) -> list[float] | None:
    text = clean(value).replace("\u00a0", " ").replace(";", " ")
    match = re.search(r"(-?\d+(?:[.,]\d+)?)\s*,\s*(-?\d+(?:[.,]\d+)?)", text)
    if not match:
        match = re.search(r"(-?\d+(?:[.,]\d+)?)\s+(-?\d+(?:[.,]\d+)?)", text)
    if not match:
        return None
    first = float(match.group(1).replace(",", "."))
    second = float(match.group(2).replace(",", "."))
    if abs(first) <= 90 and abs(second) <= 180:
        lat, lng = first, second
    elif abs(second) <= 90 and abs(first) <= 180:
        lat, lng = second, first
    else:
        return None
    return [lng, lat]


def feature(feature_id: str, coordinates: list[float], properties: dict[str, object]) -> dict:
    return {
        "type": "Feature",
        "id": feature_id,
        "properties": {key: value for key, value in properties.items() if value not in (None, "")},
        "geometry": {"type": "Point", "coordinates": coordinates},
    }


def collection(items: list[dict]) -> dict:
    return {"type": "FeatureCollection", "features": items}


def read_mno(path: Path) -> tuple[list[dict], dict]:
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True).active
    features: list[dict] = []
    total = 0
    missing = 0
    for row_number, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(clean(value) for value in row):
            continue
        total += 1
        coordinates = parse_coordinates(row[9] if len(row) > 9 else None)
        if coordinates is None:
            missing += 1
            continue
        features.append(
            feature(
                f"mno-{clean(row[8]) or clean(row[2]) or 'item'}-{row_number}",
                coordinates,
                {
                    "category": "МНО",
                    "layer": "mno",
                    "district": clean(row[1]),
                    "section": f"Участок {clean(row[4])}" if clean(row[4]) else "Без участка",
                    "name": clean(row[5]) or "Место накопления отходов",
                    "address": clean(row[5]),
                    "source_address": clean(row[3]),
                    "mno_type": clean(row[6]),
                    "container_count": clean(row[7]),
                    "id": clean(row[2]),
                    "mno_id": clean(row[8]),
                },
            )
        )
    return features, {"source_rows": total, "loaded": len(features), "missing_coordinates": missing}


def read_dp_sp(path: Path) -> tuple[dict[str, list[dict]], dict]:
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True)["Готовый файл"]
    output = {"dp": [], "sp": []}
    total = {"source_rows": 0, "loaded": 0, "missing_coordinates": 0, "other_types": {}}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(clean(value) for value in row):
            continue
        total["source_rows"] += 1
        kind = clean(row[3] if len(row) > 3 else None)
        layer = "dp" if kind == "Детская площадка" else "sp" if kind == "Спортивная площадка" else None
        if layer is None:
            total["other_types"][kind or "Без типа"] = total["other_types"].get(kind or "Без типа", 0) + 1
            continue
        coordinates = parse_coordinates(row[11] if len(row) > 11 else None)
        if coordinates is None:
            total["missing_coordinates"] += 1
            continue
        category = "ДП" if layer == "dp" else "СП"
        output[layer].append(
            feature(
                f"{layer}-{clean(row[6]) or 'item'}-{row_number}",
                coordinates,
                {
                    "category": category,
                    "layer": layer,
                    "district": clean(row[1]),
                    "section": f"Участок {clean(row[9])}" if clean(row[9]) else "Без участка",
                    "name": clean(row[8]) or kind,
                    "address": clean(row[5]) or clean(row[4]),
                    "type": kind,
                    "parent_id": clean(row[6]),
                    "parent_name": clean(row[8]),
                    "site_type": clean(row[4]),
                    "improvement_year": clean(row[10]),
                },
            )
        )
    total["loaded"] = len(output["dp"]) + len(output["sp"])
    return output, total


def write_json(path: Path, data: dict) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mno", type=Path, required=True)
    parser.add_argument("--dp-sp", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=Path("."))
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)

    mno, mno_stats = read_mno(args.mno)
    dp_sp, dp_sp_stats = read_dp_sp(args.dp_sp)
    write_json(args.output / "mno.geojson", collection(mno))
    write_json(args.output / "dp.geojson", collection(dp_sp["dp"]))
    write_json(args.output / "sp.geojson", collection(dp_sp["sp"]))
    write_json(
        args.output / "layers-manifest.json",
        {
            "updated_at": date.today().isoformat(),
            "layers": {
                "mno": {"name": "МНО", "file": "mno.geojson", **mno_stats},
                "dp": {"name": "Детские площадки", "file": "dp.geojson", "loaded": len(dp_sp["dp"])},
                "sp": {"name": "Спортивные площадки", "file": "sp.geojson", "loaded": len(dp_sp["sp"])},
            },
            "dp_sp_source": dp_sp_stats,
        },
    )
    print(json.dumps({"mno": mno_stats, "dp_sp": dp_sp_stats, "dp": len(dp_sp["dp"]), "sp": len(dp_sp["sp"])}, ensure_ascii=False))


if __name__ == "__main__":
    main()
